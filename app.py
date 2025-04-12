import requests
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# === CONFIGURAÇÕES ===
API_KEY = "7dff53a27a84f8978042b8fc510d01c9"
BASE_URL = "http://api.openweathermap.org/data/2.5/weather"

# === FUNÇÃO PARA BUSCAR DADOS ===


def buscar_dados():
    cidade = entrada_cidade.get()
    if not cidade:
        messagebox.showwarning("Aviso", "Digite uma cidade!")
        return

    params = {"q": cidade, "appid": API_KEY,
              "units": "metric", "lang": "pt_br"}
    resposta = requests.get(BASE_URL, params=params)

    if resposta.status_code == 200:
        dados = resposta.json()
        temperatura = dados["main"]["temp"]
        umidade = dados["main"]["humidity"]
        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        salvar_em_planilha(agora, cidade, temperatura, umidade)
        messagebox.showinfo(
            "Sucesso", f"Dados salvos com sucesso!\n\nCidade: {cidade}\nTemp: {temperatura}°C\nUmidade: {umidade}%")
    else:
        messagebox.showerror("Erro", f"Erro ao buscar dados para {cidade}")

# === FUNÇÃO PARA SALVAR DADOS ===


def salvar_em_planilha(data_hora, cidade, temperatura, umidade):
    nome_arquivo = "dados_climaticos.xlsx"
    if os.path.exists(nome_arquivo):
        planilha = load_workbook(nome_arquivo)
        aba = planilha.active
    else:
        planilha = Workbook()
        aba = planilha.active
        aba.append(["Data / Hora", "Cidade",
                   "Temperatura (°C)", "Umidade (%)"])

    aba.append([data_hora, cidade, temperatura, umidade])
    planilha.save(nome_arquivo)


# === INTERFACE GRÁFICA ===
janela = tk.Tk()
janela.title("Coletor de Dados Climáticos")

tk.Label(janela, text="Digite uma cidade:").pack(pady=5)
entrada_cidade = tk.Entry(janela, width=30)
entrada_cidade.pack(pady=5)

btn_buscar = tk.Button(janela, text="Buscar Previsão", command=buscar_dados)
btn_buscar.pack(pady=10)

janela.mainloop()
